from openpyxl import load_workbook
import pandas as pd

source_name = "buaa.xlsx"

destination_df = pd.DataFrame(columns=["课程名", "分数", "学分", "学时", "学时单位","课程类别","学期"])

# 定义列映射关系，将source中的列名映射到destination的列名
column_mapping = {
    "课程名称": "课程名",
    "成绩": "分数",
    "学分": "学分",
    "学时":"学时",
    "性质":"课程类别",
}

column_mapping_index = {
    0: "课程名",
    4: "分数",
    3: "学分",
    2:"学时",
    1:"课程类别",
}


# 加载工作簿和指定工作表
workbook = load_workbook(filename=source_name)
sheet = workbook["sheet1"]

def get_time(s):
    year = int(s[:4])
    season = s[4:]
    if season == "秋季":
        return "-".join([str(year), str(year+1), "1"])
    elif season == "春季":
        return "-".join([str(year-1), str(year), "2"])
    elif season == "夏季":
        return "-".join([str(year-1), str(year), "3"])


# 遍历每一行
curtime = '2023-2024-2'
for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始（如果有标题）
    # row 是一个包含这一行所有单元格数据的元组
    # print(row)
    # 处理时间
    if isinstance(row[0], str) and row[0][0].isdigit():
        curtime = get_time(row[0])
        continue
    if row[0] == '':
        continue

    dst_row = {col: None for col in destination_df.columns}
    for col_index, col_name in column_mapping_index.items():
        dst_row[col_name] = row[col_index]
    dst_row['学时单位'] = '学时'
    dst_row['学期'] = curtime
    dst_series = pd.Series(dst_row)
    destination_df = pd.concat([destination_df, pd.DataFrame([dst_row])], ignore_index=True)
print(destination_df)
destination_df.to_excel("xxw.xlsx", index=False)
